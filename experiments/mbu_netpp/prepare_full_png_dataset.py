from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook

from backend.app.utils.image_io import write_image

from experiments.mbu_netpp.common import ensure_dir, save_json
from experiments.mbu_netpp.preparation import detect_crop_height


def canonical_stem(text: str) -> str:
    stem = Path(str(text).strip()).stem
    stem = re.sub(r"\s+", "", stem)
    return stem.lower()


def read_gray(path: Path):
    image = cv2.imdecode(
        np.fromfile(str(path), dtype=np.uint8),
        cv2.IMREAD_GRAYSCALE,
    )
    if image is None:
        raise ValueError(f"无法读取图像: {path}")
    return image


def smooth_signal(values: np.ndarray, window_size: int = 9) -> np.ndarray:
    if window_size <= 1:
        return values.astype(np.float32)
    kernel = np.ones(window_size, dtype=np.float32) / float(window_size)
    return np.convolve(values.astype(np.float32), kernel, mode="same")


def detect_left_sidebar_width(image: np.ndarray) -> int:
    height, width = image.shape[:2]
    if height <= width:
        return 0

    col_mean = smooth_signal(image.mean(axis=0), window_size=11)
    col_std = smooth_signal(image.std(axis=0), window_size=11)
    search_start = max(16, int(width * 0.08))
    search_end = min(width - 20, int(width * 0.40))
    if search_end <= search_start + 10:
        return 0

    left_level = float(np.median(col_mean[:search_start]))
    content_start = min(max(search_end + 8, int(width * 0.45)), width - 24)
    content_end = min(width, content_start + max(24, int(width * 0.18)))
    content_level = float(np.median(col_mean[content_start:content_end]))
    if content_level - left_level < 35:
        return 0

    threshold = left_level + 0.55 * (content_level - left_level)
    for x in range(search_start, search_end):
        if col_mean[x] >= threshold and col_std[x] >= 18:
            return x
    return 0


def expected_crop_height(image_height: int, footer_type: str, left_crop_width: int) -> int | None:
    if left_crop_width > 0:
        return image_height
    if footer_type == "2" and image_height == 888:
        return 767
    if footer_type == "1" and image_height in {946, 948}:
        return int(round(image_height * 0.811))
    return None


def copy_xlsx_to_ascii_path(xlsx_path: Path, output_root: Path) -> Path:
    temp_dir = ensure_dir(output_root / "_tmp")
    safe_path = temp_dir / "footer_info_map.xlsx"
    shutil.copy2(xlsx_path, safe_path)
    return safe_path


def load_mapping_rows(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, str]] = []
    for raw_file, raw_type in worksheet.iter_rows(min_row=2, values_only=True):
        if not raw_file:
            continue
        rows.append(
            {
                "file": str(raw_file).strip(),
                "footer_type": str(raw_type).strip(),
            }
        )
    return rows


def build_image_lookup(images_dir: Path) -> dict[str, Path]:
    lookup: dict[str, Path] = {}
    for path in sorted(images_dir.glob("*.png")):
        lookup[canonical_stem(path.name)] = path
    return lookup


def save_manifest_xlsx(rows: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "manifest"
    headers = [
        "xlsx_file",
        "matched_png",
        "footer_type",
        "footer_pixels",
        "orig_width",
        "orig_height",
        "detected_crop_height",
        "crop_height",
        "crop_method",
        "status",
        "output_image_path",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    workbook.save(output_path)


def prepare_full_png_dataset(
    images_dir: str | Path,
    mapping_xlsx: str | Path,
    output_root: str | Path,
    footer_pixels_by_type: dict[str, int] | None = None,
    crop_detection_ratio: float = 0.72,
) -> dict[str, Any]:
    images_dir = Path(images_dir)
    mapping_xlsx = Path(mapping_xlsx)
    output_root = ensure_dir(output_root)
    cropped_images_dir = ensure_dir(output_root / "images")
    manifests_dir = ensure_dir(output_root / "manifests")

    footer_map = footer_pixels_by_type or {"1": 125, "2": 81}
    safe_xlsx_path = copy_xlsx_to_ascii_path(mapping_xlsx, output_root)
    mapping_rows = load_mapping_rows(safe_xlsx_path)
    image_lookup = build_image_lookup(images_dir)

    manifest_rows: list[dict[str, Any]] = []
    matched_keys: set[str] = set()

    for mapping_row in mapping_rows:
        raw_file = mapping_row["file"]
        footer_type = mapping_row["footer_type"]
        canonical = canonical_stem(raw_file)
        image_path = image_lookup.get(canonical)
        footer_pixels = int(footer_map.get(footer_type, 0))

        if image_path is None:
            manifest_rows.append(
                {
                    "xlsx_file": raw_file,
                    "matched_png": "",
                    "footer_type": footer_type,
                    "footer_pixels": footer_pixels,
                    "orig_width": "",
                    "orig_height": "",
                    "detected_crop_height": "",
                    "crop_height": "",
                    "crop_method": "",
                    "status": "missing_png",
                    "output_image_path": "",
                }
            )
            continue

        matched_keys.add(canonical)
        image = read_gray(image_path)
        orig_height, orig_width = image.shape[:2]
        left_crop_width = int(detect_left_sidebar_width(image))
        if left_crop_width > 0:
            image = image[:, left_crop_width:]
        height, width = image.shape[:2]

        expected_height = expected_crop_height(height, footer_type, left_crop_width)
        fallback_crop_height = max(1, height - footer_pixels) if left_crop_width == 0 else height
        detected_crop_height = int(detect_crop_height(image, start_ratio=crop_detection_ratio)) if left_crop_width == 0 else height

        if expected_height is not None:
            if left_crop_width > 0:
                crop_height = expected_height
                crop_method = "left_sidebar_template"
            elif abs(detected_crop_height - expected_height) <= 8:
                crop_height = detected_crop_height
                crop_method = "auto_detect"
            else:
                crop_height = expected_height
                crop_method = "expected_template"
        elif left_crop_width > 0:
            crop_height = height
            crop_method = "left_sidebar"
        elif height - 260 <= detected_crop_height <= height - 40:
            crop_height = detected_crop_height
            crop_method = "auto_detect"
        else:
            crop_height = fallback_crop_height
            crop_method = "type_fallback"
        cropped = image[:crop_height, :]
        output_image_path = cropped_images_dir / image_path.name
        write_image(output_image_path, cropped)

        manifest_rows.append(
            {
                    "xlsx_file": raw_file,
                    "matched_png": image_path.name,
                    "footer_type": footer_type,
                    "footer_pixels": footer_pixels,
                    "orig_width": orig_width,
                    "orig_height": orig_height,
                    "detected_crop_height": detected_crop_height,
                    "crop_height": crop_height,
                    "crop_method": crop_method,
                    "status": "ok",
                "output_image_path": str(output_image_path),
            }
        )

    for canonical, image_path in sorted(image_lookup.items()):
        if canonical in matched_keys:
            continue
        manifest_rows.append(
            {
                "xlsx_file": "",
                "matched_png": image_path.name,
                "footer_type": "",
                "footer_pixels": "",
                "orig_width": "",
                "orig_height": "",
                "detected_crop_height": "",
                "crop_height": "",
                "crop_method": "",
                "status": "missing_xlsx_row",
                "output_image_path": "",
            }
        )

    csv_path = manifests_dir / "manifest.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "xlsx_file",
                "matched_png",
                "footer_type",
                "footer_pixels",
                "orig_width",
                "orig_height",
                "detected_crop_height",
                "crop_height",
                "crop_method",
                "status",
                "output_image_path",
            ],
        )
        writer.writeheader()
        for row in manifest_rows:
            writer.writerow(row)

    save_manifest_xlsx(manifest_rows, manifests_dir / "manifest.xlsx")
    save_json(
        manifests_dir / "dataset.json",
        {
            "images_dir": str(images_dir),
            "mapping_xlsx": str(mapping_xlsx),
            "footer_pixels_by_type": footer_map,
            "crop_detection_ratio": crop_detection_ratio,
            "num_rows": len(mapping_rows),
            "num_processed": sum(1 for row in manifest_rows if row["status"] == "ok"),
            "num_missing_png": sum(1 for row in manifest_rows if row["status"] == "missing_png"),
            "num_missing_xlsx_row": sum(1 for row in manifest_rows if row["status"] == "missing_xlsx_row"),
            "items": manifest_rows,
        },
    )
    return {
        "output_root": str(output_root),
        "num_processed": sum(1 for row in manifest_rows if row["status"] == "ok"),
        "num_missing_png": sum(1 for row in manifest_rows if row["status"] == "missing_png"),
        "num_missing_xlsx_row": sum(1 for row in manifest_rows if row["status"] == "missing_xlsx_row"),
        "manifest_csv": str(csv_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按 xlsx 类型裁掉 full_png 底部信息栏")
    parser.add_argument("--images-dir", required=True, help="full_png 目录")
    parser.add_argument("--mapping-xlsx", required=True, help="文件与底部信息类型映射 xlsx")
    parser.add_argument("--output-root", required=True, help="裁切后数据集输出目录")
    parser.add_argument("--type1-footer-px", type=int, default=125, help="type 1 底栏像素高度")
    parser.add_argument("--type2-footer-px", type=int, default=81, help="type 2 底栏像素高度")
    parser.add_argument("--crop-detection-ratio", type=float, default=0.72, help="自动检测底栏起始位置的起始比例")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = prepare_full_png_dataset(
        images_dir=args.images_dir,
        mapping_xlsx=args.mapping_xlsx,
        output_root=args.output_root,
        footer_pixels_by_type={
            "1": int(args.type1_footer_px),
            "2": int(args.type2_footer_px),
        },
        crop_detection_ratio=float(args.crop_detection_ratio),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
